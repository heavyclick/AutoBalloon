"""
Export Service - AS9102 Rev C Compliant
Generates CSV and AS9102 Form 3 Excel exports for First Article Inspection.

Updated to match official AS9102 Rev C template format with:
- Proper header section (Part Number, Part Name, Serial Number, FAI Identifier)
- Official column structure matching aerospace standards
- Multi-page support with Sheet column
- Grid detection status indicator
"""
import csv
import io
from datetime import datetime
from typing import Optional, List, Dict, Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.utils import get_column_letter

from models import ExportFormat, ExportTemplate, ExportMetadata


class ExportService:
    """
    Generates AS9102 Rev C compliant export files for FAI inspection data.
    """
    
    # ==================
    # Color Scheme (Professional aerospace look)
    # ==================
    HEADER_BG = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")  # Dark blue
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=10, name="Arial")
    SUBHEADER_BG = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")  # Light gray-blue
    SUBHEADER_FONT = Font(bold=True, color="000000", size=9, name="Arial")
    
    TITLE_FONT = Font(bold=True, color="000000", size=14, name="Arial")
    SUBTITLE_FONT = Font(bold=True, color="1F4E79", size=11, name="Arial")
    LABEL_FONT = Font(bold=True, color="000000", size=9, name="Arial")
    DATA_FONT = Font(color="000000", size=9, name="Arial")
    NOTE_FONT = Font(italic=True, color="666666", size=8, name="Arial")
    
    # Borders
    THIN_BORDER = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    MEDIUM_BORDER = Border(
        left=Side(style='medium', color='000000'),
        right=Side(style='medium', color='000000'),
        top=Side(style='medium', color='000000'),
        bottom=Side(style='medium', color='000000')
    )
    
    # Alignments
    CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
    LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)
    RIGHT = Alignment(horizontal='right', vertical='center')
    
    # AS9102 Form 3 Column Headers (official names)
    FORM3_HEADERS = [
        ("5. Char\nNo.", 8),           # Column A - width 8
        ("6. Reference\nLocation", 12), # Column B - width 12
        ("7. Characteristic\nDesignator", 14),  # Column C - width 14
        ("8. Requirement", 25),         # Column D - width 25
        ("9. Results", 15),             # Column E - width 15
        ("10. Designed/\nQualified\nTooling", 12),  # Column F - width 12
        ("11. Non-\nConformance\nNumber", 12),  # Column G - width 12
        ("Sheet", 8),                   # Column H - width 8 (for multi-page)
    ]
    
    def generate_export(
        self,
        dimensions: List[Dict],
        format: ExportFormat,
        template: ExportTemplate = ExportTemplate.AS9102_FORM3,
        metadata: Optional[ExportMetadata] = None,
        filename: str = "inspection",
        grid_detected: bool = True,
        total_pages: int = 1
    ) -> tuple[bytes, str, str]:
        """
        Generate export file.
        
        Args:
            dimensions: List of dimension dicts with id, value, zone, page (optional)
            format: Export format (CSV or XLSX)
            template: Export template (SIMPLE or AS9102_FORM3)
            metadata: Optional metadata (part number, revision, etc.)
            filename: Base filename (without extension)
            grid_detected: Whether grid was auto-detected or using standard
            total_pages: Total number of pages in the drawing
            
        Returns:
            Tuple of (file_bytes, content_type, full_filename)
        """
        if format == ExportFormat.CSV:
            return self._generate_csv(dimensions, filename, total_pages)
        else:
            if template == ExportTemplate.SIMPLE:
                return self._generate_simple_xlsx(dimensions, filename, total_pages)
            else:
                return self._generate_as9102_xlsx(dimensions, metadata, filename, grid_detected, total_pages)
    
    def _generate_csv(
        self, 
        dimensions: List[Dict],
        filename: str,
        total_pages: int = 1
    ) -> tuple[bytes, str, str]:
        """Generate simple CSV export"""
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Header - include Sheet column if multi-page
        if total_pages > 1:
            writer.writerow(["Char No", "Reference Location", "Requirement", "Sheet"])
        else:
            writer.writerow(["Char No", "Reference Location", "Requirement"])
        
        # Data
        for dim in dimensions:
            row = [
                dim.get("id", ""),
                dim.get("zone", "—"),
                dim.get("value", "")
            ]
            if total_pages > 1:
                row.append(dim.get("page", 1))
            writer.writerow(row)
        
        csv_bytes = output.getvalue().encode('utf-8')
        return (
            csv_bytes,
            "text/csv",
            f"{filename}_inspection.csv"
        )
    
    def _generate_simple_xlsx(
        self,
        dimensions: List[Dict],
        filename: str,
        total_pages: int = 1
    ) -> tuple[bytes, str, str]:
        """Generate simple Excel export with basic columns"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Inspection"
        
        # Headers
        headers = ["Char No", "Reference Location", "Requirement"]
        if total_pages > 1:
            headers.append("Sheet")
            
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_BG
            cell.alignment = self.CENTER
            cell.border = self.THIN_BORDER
        
        # Data
        for row_idx, dim in enumerate(dimensions, start=2):
            ws.cell(row=row_idx, column=1, value=dim.get("id", "")).border = self.THIN_BORDER
            ws.cell(row=row_idx, column=2, value=dim.get("zone", "—")).border = self.THIN_BORDER
            ws.cell(row=row_idx, column=3, value=dim.get("value", "")).border = self.THIN_BORDER
            if total_pages > 1:
                ws.cell(row=row_idx, column=4, value=dim.get("page", 1)).border = self.THIN_BORDER
        
        # Column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 25
        if total_pages > 1:
            ws.column_dimensions['D'].width = 8
        
        output = io.BytesIO()
        wb.save(output)
        return (output.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", f"{filename}_inspection.xlsx")
    
    def _generate_as9102_xlsx(
        self,
        dimensions: List[Dict],
        metadata: Optional[ExportMetadata],
        filename: str,
        grid_detected: bool = True,
        total_pages: int = 1
    ) -> tuple[bytes, str, str]:
        """
        Generate AS9102 Rev C Form 3 compliant Excel export.
        
        Structure:
        - Row 1: Title
        - Row 2: Subtitle
        - Row 3: Empty
        - Row 4-5: Part info header row 1 (labels + values)
        - Row 6: Empty
        - Row 7: Column headers (Characteristic Accountability | Inspection Results)
        - Row 8: Column sub-headers
        - Row 9+: Data rows
        - Footer: Grid detection note, signature lines
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "AS9102 Form 3"
        
        # Build filename from metadata
        parts = []
        if metadata:
            if metadata.part_number:
                parts.append(metadata.part_number)
            if metadata.revision:
                parts.append(f"Rev{metadata.revision}")
        if not parts:
            parts.append(filename)
        parts.append("FAI_Form3")
        full_filename = "_".join(parts) + ".xlsx"
        
        current_row = 1
        
        # ==================
        # TITLE SECTION
        # ==================
        ws.merge_cells('A1:H1')
        title_cell = ws.cell(row=1, column=1, value="AS9102 First Article Inspection")
        title_cell.font = self.TITLE_FONT
        title_cell.alignment = self.CENTER
        
        ws.merge_cells('A2:H2')
        subtitle_cell = ws.cell(row=2, column=1, value="Form 3: Characteristic Accountability, Verification and Compatibility Evaluation")
        subtitle_cell.font = self.SUBTITLE_FONT
        subtitle_cell.alignment = self.CENTER
        
        current_row = 4
        
        # ==================
        # PART INFO SECTION
        # ==================
        # Row 4: Labels
        info_labels = [
            ("1. Part Number:", "A4"),
            ("2. Part Name:", "C4"),
            ("3. Serial Number:", "E4"),
            ("4. FAI Identifier:", "G4"),
        ]
        
        for label, cell_ref in info_labels:
            cell = ws[cell_ref]
            cell.value = label
            cell.font = self.LABEL_FONT
            cell.alignment = self.LEFT
        
        # Row 5: Values
        info_values = [
            (metadata.part_number if metadata else "", "A5", "B5"),
            (metadata.part_name if metadata else "", "C5", "D5"),
            ("", "E5", "F5"),  # Serial Number - blank for user
            ("", "G5", "H5"),  # FAI Identifier - blank for user
        ]
        
        for value, start_cell, end_cell in info_values:
            ws.merge_cells(f"{start_cell}:{end_cell}")
            cell = ws[start_cell]
            cell.value = value
            cell.font = self.DATA_FONT
            cell.alignment = self.LEFT
            cell.border = self.THIN_BORDER
            # Also border the end cell
            ws[end_cell].border = self.THIN_BORDER
        
        current_row = 7
        
        # ==================
        # SECTION HEADERS
        # ==================
        # Row 7: Section headers
        ws.merge_cells('A7:D7')
        char_header = ws.cell(row=7, column=1, value="Characteristic Accountability:")
        char_header.font = self.SUBHEADER_FONT
        char_header.fill = self.SUBHEADER_BG
        char_header.alignment = self.CENTER
        char_header.border = self.THIN_BORDER
        
        ws.merge_cells('E7:G7')
        results_header = ws.cell(row=7, column=5, value="Inspection / Test Results:")
        results_header.font = self.SUBHEADER_FONT
        results_header.fill = self.SUBHEADER_BG
        results_header.alignment = self.CENTER
        results_header.border = self.THIN_BORDER
        
        # Sheet header (column H)
        if total_pages > 1:
            sheet_header = ws.cell(row=7, column=8, value="")
            sheet_header.fill = self.SUBHEADER_BG
            sheet_header.border = self.THIN_BORDER
        
        current_row = 8
        
        # ==================
        # COLUMN HEADERS (Row 8)
        # ==================
        for col_idx, (header, width) in enumerate(self.FORM3_HEADERS, start=1):
            # Skip Sheet column if single page
            if col_idx == 8 and total_pages <= 1:
                continue
                
            cell = ws.cell(row=8, column=col_idx, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_BG
            cell.alignment = self.CENTER
            cell.border = self.THIN_BORDER
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        
        current_row = 9
        
        # ==================
        # DATA ROWS
        # ==================
        for dim in dimensions:
            # Column A: Char No
            cell_a = ws.cell(row=current_row, column=1, value=dim.get("id", ""))
            cell_a.font = self.DATA_FONT
            cell_a.alignment = self.CENTER
            cell_a.border = self.THIN_BORDER
            
            # Column B: Reference Location (Zone)
            zone = dim.get("zone", "")
            cell_b = ws.cell(row=current_row, column=2, value=zone if zone else "—")
            cell_b.font = self.DATA_FONT
            cell_b.alignment = self.CENTER
            cell_b.border = self.THIN_BORDER
            
            # Column C: Characteristic Designator (blank - user fills)
            cell_c = ws.cell(row=current_row, column=3, value="")
            cell_c.border = self.THIN_BORDER
            
            # Column D: Requirement (dimension value)
            cell_d = ws.cell(row=current_row, column=4, value=dim.get("value", ""))
            cell_d.font = self.DATA_FONT
            cell_d.alignment = self.LEFT
            cell_d.border = self.THIN_BORDER
            
            # Column E: Results (blank - inspector fills, or from CMM import)
            actual = dim.get("actual", "")
            cell_e = ws.cell(row=current_row, column=5, value=actual)
            cell_e.font = self.DATA_FONT
            cell_e.alignment = self.CENTER
            cell_e.border = self.THIN_BORDER
            
            # Column F: Designed/Qualified Tooling (blank - inspector fills)
            cell_f = ws.cell(row=current_row, column=6, value="")
            cell_f.border = self.THIN_BORDER
            
            # Column G: Non-Conformance Number (blank - inspector fills)
            cell_g = ws.cell(row=current_row, column=7, value="")
            cell_g.border = self.THIN_BORDER
            
            # Column H: Sheet (page number) - only if multi-page
            if total_pages > 1:
                cell_h = ws.cell(row=current_row, column=8, value=dim.get("page", 1))
                cell_h.font = self.DATA_FONT
                cell_h.alignment = self.CENTER
                cell_h.border = self.THIN_BORDER
            
            current_row += 1
        
        # Add some empty rows for manual entries
        for _ in range(5):
            for col in range(1, 8 if total_pages <= 1 else 9):
                cell = ws.cell(row=current_row, column=col, value="")
                cell.border = self.THIN_BORDER
            current_row += 1
        
        current_row += 1  # Empty row
        
        # ==================
        # FOOTER SECTION
        # ==================
        
        # Grid detection note
        if not grid_detected:
            ws.merge_cells(f'A{current_row}:H{current_row}')
            note_cell = ws.cell(row=current_row, column=1, 
                value="Note: Grid zones calculated using standard 8×4 grid (H-A × 4-1). Drawing grid was not auto-detected.")
            note_cell.font = self.NOTE_FONT
            note_cell.alignment = self.LEFT
            current_row += 2
        
        # Generation info
        ws.merge_cells(f'A{current_row}:H{current_row}')
        gen_cell = ws.cell(row=current_row, column=1, 
            value=f"Generated by AutoBalloon on {datetime.now().strftime('%Y-%m-%d %H:%M')} | {len(dimensions)} characteristics | {total_pages} sheet(s)")
        gen_cell.font = self.NOTE_FONT
        gen_cell.alignment = self.LEFT
        current_row += 2
        
        # Signature lines
        ws.cell(row=current_row, column=1, value="Prepared By:").font = self.LABEL_FONT
        ws.cell(row=current_row, column=3, value="Date:").font = self.LABEL_FONT
        ws.cell(row=current_row, column=5, value="Approved By:").font = self.LABEL_FONT
        ws.cell(row=current_row, column=7, value="Date:").font = self.LABEL_FONT
        
        current_row += 1
        # Signature lines (underscores)
        for col in [1, 2, 3, 4, 5, 6, 7, 8]:
            cell = ws.cell(row=current_row, column=col, value="")
            cell.border = Border(bottom=Side(style='thin', color='000000'))
        
        # ==================
        # PAGE SETUP
        # ==================
        ws.page_setup.orientation = 'landscape'
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        
        # Print titles (repeat header rows on each page)
        ws.print_title_rows = '1:8'
        
        # Freeze panes (keep headers visible when scrolling)
        ws.freeze_panes = 'A9'
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        xlsx_bytes = output.getvalue()
        
        return (
            xlsx_bytes,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            full_filename
        )
    
    def generate_multi_page_export(
        self,
        pages_data: List[Dict],
        format: ExportFormat,
        template: ExportTemplate = ExportTemplate.AS9102_FORM3,
        metadata: Optional[ExportMetadata] = None,
        filename: str = "inspection",
        grid_statuses: Optional[List[bool]] = None
    ) -> tuple[bytes, str, str]:
        """
        Generate export for multi-page drawings.
        
        Args:
            pages_data: List of page dicts, each containing 'page_number' and 'dimensions'
            format: Export format
            template: Export template
            metadata: Optional metadata
            filename: Base filename
            grid_statuses: List of booleans indicating if grid was detected per page
            
        Returns:
            Tuple of (file_bytes, content_type, full_filename)
        """
        # Flatten all dimensions with page numbers
        all_dimensions = []
        for page_data in pages_data:
            page_num = page_data.get('page_number', 1)
            for dim in page_data.get('dimensions', []):
                dim_copy = dim.copy()
                dim_copy['page'] = page_num
                all_dimensions.append(dim_copy)
        
        # Determine if any grid was not detected
        grid_detected = True
        if grid_statuses:
            grid_detected = all(grid_statuses)
        
        total_pages = len(pages_data)
        
        return self.generate_export(
            dimensions=all_dimensions,
            format=format,
            template=template,
            metadata=metadata,
            filename=filename,
            grid_detected=grid_detected,
            total_pages=total_pages
        )


# Singleton instance
export_service = ExportService()
